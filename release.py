# -*- coding = utf-8 -*-
# @Time :2021/6/13 1:27 下午
# @Author: XZL
# @File : release.py
# @Software: PyCharm
import cv2 as cv
import numpy as np
import platform
import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, colors

separator = '/' if platform.system() == 'Darwin' else '\\'
dir_path = os.getcwd()
# 创建图片导出文件夹
save_pic_dir = dir_path + separator + 'export'
if not os.path.exists(save_pic_dir):
    os.mkdir(save_pic_dir)

# 全局维护的data，index为当前下标
save_data = {
    'index': 0,
    'length': 0,
    'picFullPath': [],
    'picPrefix': [],
    'area': [],
    'areaRate': [],
    'whiteArea': [],
    'whiteAreaRate': []
}


# 获取文件保存路径 bin：二值图 con：凸包图
def save_pic_path(type):
    root_path = save_pic_dir + separator
    if type == 'bin':
        return root_path + 'bin_' + save_data.get('picPrefix')[save_data.get('index')] + '.jpg'
    elif type == 'con':
        return root_path + 'con_' + save_data.get('picPrefix')[save_data.get('index')] + '.jpg'


def dataloader(path):
    # 读取文件夹下所有文件
    file_list = os.listdir(path)
    for item in file_list:
        split_name = os.path.splitext(item)
        # 读取指定类型文件
        if split_name[1] in ['.bmp', '.png', '.jpg', '.jpeg']:
            single_img_path = path + separator + item
            save_data.get('picFullPath').append(single_img_path)
            save_data.get('picPrefix').append(split_name[0])
    save_data['length'] = len(save_data.get('picPrefix'))

    # print(single_img_path)
    # print(cv.imread(single_img_path).shape)


def threshold_demo(img):
    gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    # 全局阈值二值化 主要用：cv.THRESH_OTSU和cv.THRESH_TRIANGLE两种
    ret, binary = cv.threshold(gray, 0, 255, cv.THRESH_BINARY + cv.THRESH_OTSU)
    # cv.imshow('binary_otsu', binary)
    cv.imwrite(save_pic_path('bin'), binary, [int(cv.IMWRITE_JPEG_QUALITY), 100])
    return binary


def canny(image):
    blur = cv.GaussianBlur(image, (3, 3), 0)  # 高斯模糊去噪声
    # dege 后面两个比值一般是3：1或者2：1
    blue_c, green_c, red_c = cv.split(blur)
    edge_canny_one = cv.Canny(green_c, 30, 150)
    return edge_canny_one


def outter_contours(img):
    """
    划定图像的外接凸包，计算图像的总面积
    """
    dst = canny(img)
    t_img = img.copy()
    cnt = np.argwhere(dst >= 1)
    # 交换x:y满足opencv绘图规则
    cnt = cnt[:, [1, 0]]
    hull = cv.convexHull(cnt)
    # 只有当有凸包时才记录
    if hull is not None:
        area = cv.contourArea(hull)
        imgRet = cv.polylines(t_img, [hull], True, (0, 0, 255), 1)  # 绘制外边框多边形
        save_data.get('area').append(str(area))
        save_data.get('areaRate').append(str(round(area * 100 / (img.shape[0] * img.shape[1]), 2)) + '%')

        print('物件面积:' + str(area) + 'px')
        print('物件占比:' + str(round(area * 100 / (img.shape[0] * img.shape[1]), 2)))
        # 100质量保存凸包图片
        cv.imwrite(save_pic_path('con'), imgRet, [int(cv.IMWRITE_JPEG_QUALITY), 100])
        # cv.imshow('outer', imgRet)
    else:
        area = 0
        save_data.get('area').append('None')
        save_data.get('areaRate').append('None')
    return area


def count_white_area(img):
    """
    统计所有白色像素点数量
    """
    binary = threshold_demo(img)
    return np.argwhere(binary == 255).shape[0]


def pipline(index):
    print('>>>>>> ' + save_data.get('picPrefix')[index] + ' starting' + ' <<<<<<')
    argv = save_data.get('picFullPath')[index]
    img = cv.imread(argv)
    h = img.shape[0]
    w = img.shape[1]
    img[h - 50:h, w - 50:w, :] = 0
    img[h - 50:h, :50, :] = 0
    # img = cv.imshow(argv)
    # 取绿色通道
    # cv.imshow('input', img)

    sum_area = outter_contours(img)
    if sum_area > 0:
        white_area = count_white_area(img)
        save_data.get('whiteArea').append(str(white_area))
        save_data.get('whiteAreaRate').append(str(round(white_area * 100 / sum_area, 2)))
        print('白色面积:' + str(white_area) + 'px')
        print('占空比:' + str(round(white_area * 100 / sum_area, 2)) + '%')

        #
        # if cv.waitKey(0) == ord('q'):
        #     cv.destroyAllWindows()
    else:
        save_data.get('whiteArea').append('None')
        save_data.get('whiteAreaRate').append('None')


def excel_handle():
    wb = Workbook()
    ws = wb.active
    ws.append(['图片名', '物件面积 px', '物件占比 %', '白色面积 px', '占空比 %'])
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 15
    warning = PatternFill("solid", fgColor='FF8C00')
    for i in range(2, save_data.get('length') + 2):
        head = ws.cell(row=i, column=1, value=save_data.get('picPrefix')[i - 2])
        _ = ws.cell(row=i, column=2, value=save_data.get('area')[i - 2])
        _ = ws.cell(row=i, column=3, value=save_data.get('areaRate')[i - 2])
        _ = ws.cell(row=i, column=4, value=save_data.get('whiteArea')[i - 2])
        _ = ws.cell(row=i, column=5, value=save_data.get('whiteAreaRate')[i - 2])
        if save_data.get('area')[i - 2] == 'None':
            head.fill = warning
    wb.save('result.xlsx')


if __name__ == '__main__':
    path = input('绝对路径: ')
    print(path)
    dataloader(path)
    for i in range(save_data.get('length')):
        save_data['index'] = i
        pipline(i)
    # print(save_data)
    excel_handle()
